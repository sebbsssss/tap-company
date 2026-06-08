"""
Excel-based occupancy data source.

Reads occupancy_monthly_CL_2025.xlsx and serves the 5 API endpoints.
This replaces the CRM /com/report/* proxy which is permanently broken.

Sheet structure (confirmed by tech, June 2026):
  Master Cleaned  — 12,294 rows, one row per (month, property, unit)
  Non-Operational — 44 non-operational windows
  Lease Start     — 14 lease-start ramp windows

Occupancy formula: sum(Total Occupied Rooms-Days) / sum(Total Rooms-Days)
This is weighted occupancy, not an average of per-unit rates.

Usage:
  bundle = load_excel("/data/occupancy.xlsx")
  # bundle is None if the file doesn't exist — callers fall back to CRM
"""

from __future__ import annotations

import calendar
import json
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


def _log(level: str, msg: str, **kwargs: Any) -> None:
    print(json.dumps({"level": level, "msg": msg, **kwargs,
                      "ts": datetime.utcnow().isoformat() + "Z"}), file=sys.stderr)


# ---------------------------------------------------------------------------
# Internal row types
# ---------------------------------------------------------------------------

@dataclass
class MasterRow:
    month: str            # "YYYY-MM"
    property_name: str
    unit: str
    room_no: str
    property_rental_type: str
    lease_start: Optional[date]
    lease_end: Optional[date]
    occ_rooms: float      # Total Occupied Rooms
    occ_room_days: float  # Total Occupied Rooms-Days
    total_rooms: float    # Total Rooms
    total_room_days: float  # Total Rooms-Days
    occupied_rate: float  # pre-computed in Excel (for cross-check only)
    quarter: str


@dataclass
class NonOpRow:
    property_name: str
    unit: str
    start_date: Optional[date]
    end_date: Optional[date]
    reason: str
    raw: dict = field(default_factory=dict)


@dataclass
class LeaseStartRow:
    property_name: str
    unit: str
    lease_start_date: Optional[date]
    tenant: str
    raw: dict = field(default_factory=dict)


@dataclass
class ExcelBundle:
    master_rows: list[MasterRow]
    non_op_rows: list[NonOpRow]
    lease_start_rows: list[LeaseStartRow]
    source_path: str
    loaded_at: str

    # Derived: {month -> {property_name -> [MasterRow]}}
    _index: dict[str, dict[str, list[MasterRow]]] = field(default_factory=dict, repr=False)

    def __post_init__(self) -> None:
        idx: dict[str, dict[str, list[MasterRow]]] = {}
        for row in self.master_rows:
            idx.setdefault(row.month, {}).setdefault(row.property_name, []).append(row)
        self._index = idx

    def months(self) -> list[str]:
        return sorted(self._index.keys())

    def property_names_for_month(self, month: str) -> list[str]:
        return sorted(self._index.get(month, {}).keys())

    def rows_for(self, month: str, property_name: Optional[str] = None) -> list[MasterRow]:
        month_data = self._index.get(month, {})
        if property_name:
            return month_data.get(property_name, [])
        result = []
        for rows in month_data.values():
            result.extend(rows)
        return result


# ---------------------------------------------------------------------------
# Month/Year parsing
# ---------------------------------------------------------------------------

_MONTH_ABBR = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    "january": 1, "february": 2, "march": 3, "april": 4, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10,
    "november": 11, "december": 12,
}


def _parse_month(val: object) -> Optional[str]:
    """
    Parse a Month/Year cell value to "YYYY-MM".

    Handles:
      - datetime / date objects  → use year + month
      - "Jan-25", "Jan 25"       → 2025-01
      - "January 2025"           → 2025-01
      - "2025-01"                → 2025-01
      - "01/2025"                → 2025-01
    """
    if val is None:
        return None
    if isinstance(val, (datetime,)):
        return f"{val.year}-{val.month:02d}"
    if isinstance(val, date):
        return f"{val.year}-{val.month:02d}"
    s = str(val).strip()
    if not s:
        return None

    # "YYYY-MM"
    if len(s) == 7 and s[4] == "-" and s[:4].isdigit() and s[5:].isdigit():
        return s

    # "MM/YYYY" or "YYYY/MM"
    if "/" in s:
        parts = s.split("/")
        if len(parts) == 2:
            a, b = parts
            if len(a) == 4 and a.isdigit():
                return f"{a}-{int(b):02d}"
            if len(b) == 4 and b.isdigit():
                return f"{b}-{int(a):02d}"

    # "Mon-YY" or "Mon-YYYY" or "Mon YY" or "Mon YYYY"
    for sep in ("-", " "):
        if sep in s:
            parts = s.split(sep, 1)
            if len(parts) == 2:
                month_part = parts[0].strip().lower()
                year_part = parts[1].strip()
                mo = _MONTH_ABBR.get(month_part)
                if mo and year_part.isdigit():
                    yr = int(year_part)
                    if yr < 100:
                        yr += 2000
                    return f"{yr}-{mo:02d}"

    # "Month YYYY" (e.g. "January 2025") — already handled by space sep above
    # but try full reverse too: "2025 January"
    parts = s.split()
    if len(parts) == 2:
        for a, b in [(parts[0], parts[1]), (parts[1], parts[0])]:
            mo = _MONTH_ABBR.get(a.lower())
            if mo and b.isdigit():
                yr = int(b)
                if yr < 100:
                    yr += 2000
                return f"{yr}-{mo:02d}"

    return None


def _parse_date_cell(val: object) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _float(val: object, default: float = 0.0) -> float:
    if val is None:
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


# ---------------------------------------------------------------------------
# Column detection
# ---------------------------------------------------------------------------

def _find_header_row(ws: "openpyxl.worksheet.worksheet.Worksheet") -> Optional[int]:
    """Return the 1-based row index of the header row (first row with multi-column text content)."""
    for row_idx in range(1, min(10, ws.max_row + 1)):
        row = ws[row_idx]
        non_empty = sum(1 for c in row if c.value is not None and str(c.value).strip())
        if non_empty >= 5:
            return row_idx
    return None


def _map_columns(header_row: list) -> dict[str, int]:
    """
    Map normalised column name → 0-based column index.
    Case-insensitive, strips punctuation/spaces.
    """
    def norm(s: str) -> str:
        return "".join(c.lower() for c in s if c.isalnum())

    # Canonical column names → possible normalised forms
    _ALIASES: dict[str, list[str]] = {
        "month":             ["monthyear", "month", "monthyr", "period"],
        "property_name":     ["propertyname", "property", "propertyname"],
        "unit":              ["unit", "unitname", "unitid"],
        "room_no":           ["roomno", "room", "roomnumber", "roomno"],
        "rental_type":       ["propertyrentaltype", "rentaltype", "type", "leasetype"],
        "lease_start":       ["rentalleasestart", "leasestart", "leasestardate", "rentalleasestardate", "leasestartdate"],
        "lease_end":         ["rentalleaseend", "leaseend", "leaseenddate", "rentalleaseenddate"],
        "occ_rooms":         ["totaloccupiedrooms"],
        "occ_room_days":     ["totaloccupiedroomsdays", "totaloccupiedroommdays", "occdays"],
        "total_rooms":       ["totalrooms"],
        "total_room_days":   ["totalroomsdays", "totalroomdays", "totalroomdays"],
        "occupied_rate":     ["occupiedrate", "occrate", "rate", "occupancyrate"],
        "booking_id_list":   ["bookingidlist", "bookingids"],
        "quarter":           ["quarter", "qtr"],
    }

    result: dict[str, int] = {}
    for col_idx, cell in enumerate(header_row):
        if cell is None or cell.value is None:
            continue
        n = norm(str(cell.value))
        for field_name, aliases in _ALIASES.items():
            if n in aliases or any(a in n for a in aliases):
                if field_name not in result:
                    result[field_name] = col_idx
    return result


# ---------------------------------------------------------------------------
# Sheet parsers
# ---------------------------------------------------------------------------

def _parse_master_cleaned(ws: "openpyxl.worksheet.worksheet.Worksheet") -> list[MasterRow]:
    header_row_idx = _find_header_row(ws)
    if header_row_idx is None:
        _log("warn", "excel_source: could not find header row in Master Cleaned")
        return []

    header = list(ws[header_row_idx])
    cols = _map_columns(header)
    _log("info", "excel_source: Master Cleaned columns mapped",
         found=list(cols.keys()), total_rows=ws.max_row)

    rows: list[MasterRow] = []
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        row = ws[row_idx]
        vals = [c.value for c in row]

        def get(field_name: str, default: Any = None) -> Any:
            idx = cols.get(field_name)
            if idx is None or idx >= len(vals):
                return default
            return vals[idx]

        month_raw = get("month")
        month = _parse_month(month_raw)
        if not month:
            continue

        prop_name = str(get("property_name", "")).strip()
        if not prop_name:
            continue

        occ_room_days = _float(get("occ_room_days"))
        total_room_days = _float(get("total_room_days"))
        if total_room_days == 0:
            continue  # skip header repetitions or blank rows

        rows.append(MasterRow(
            month=month,
            property_name=prop_name,
            unit=str(get("unit", "")).strip(),
            room_no=str(get("room_no", "")).strip(),
            property_rental_type=str(get("rental_type", "")).strip(),
            lease_start=_parse_date_cell(get("lease_start")),
            lease_end=_parse_date_cell(get("lease_end")),
            occ_rooms=_float(get("occ_rooms")),
            occ_room_days=occ_room_days,
            total_rooms=_float(get("total_rooms")),
            total_room_days=total_room_days,
            occupied_rate=_float(get("occupied_rate")),
            quarter=str(get("quarter", "")).strip(),
        ))

    _log("info", f"excel_source: parsed {len(rows)} rows from Master Cleaned")
    return rows


def _parse_generic_sheet(ws: "openpyxl.worksheet.worksheet.Worksheet") -> list[dict]:
    """Parse any sheet as a list of dicts keyed by header row values."""
    header_row_idx = _find_header_row(ws)
    if header_row_idx is None:
        return []
    header = [str(c.value).strip() if c.value else f"col_{i}"
              for i, c in enumerate(ws[header_row_idx])]
    rows = []
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        row = ws[row_idx]
        vals = [c.value for c in row]
        if not any(v is not None and str(v).strip() for v in vals):
            continue
        rows.append({header[i]: vals[i] for i in range(min(len(header), len(vals)))})
    return rows


def _parse_non_operational(ws: "openpyxl.worksheet.worksheet.Worksheet") -> list[NonOpRow]:
    raw_rows = _parse_generic_sheet(ws)
    result = []
    for r in raw_rows:
        # Try to extract property, unit, dates from whatever columns exist
        prop = ""
        unit = ""
        start = None
        end = None
        reason = ""
        for k, v in r.items():
            kn = "".join(c.lower() for c in k if c.isalnum())
            if "property" in kn and not prop:
                prop = str(v or "").strip()
            elif "unit" in kn and not unit:
                unit = str(v or "").strip()
            elif "start" in kn and not start:
                start = _parse_date_cell(v)
            elif "end" in kn and not end:
                end = _parse_date_cell(v)
            elif "reason" in kn or "note" in kn or "remark" in kn:
                reason = str(v or "").strip()
        if prop:
            result.append(NonOpRow(
                property_name=prop, unit=unit,
                start_date=start, end_date=end,
                reason=reason, raw=r,
            ))
    _log("info", f"excel_source: parsed {len(result)} Non-Operational rows")
    return result


def _parse_lease_start(ws: "openpyxl.worksheet.worksheet.Worksheet") -> list[LeaseStartRow]:
    raw_rows = _parse_generic_sheet(ws)
    result = []
    for r in raw_rows:
        prop = ""
        unit = ""
        ls_date = None
        tenant = ""
        for k, v in r.items():
            kn = "".join(c.lower() for c in k if c.isalnum())
            if "property" in kn and not prop:
                prop = str(v or "").strip()
            elif "unit" in kn and not unit:
                unit = str(v or "").strip()
            elif "start" in kn and not ls_date:
                ls_date = _parse_date_cell(v)
            elif "tenant" in kn or "member" in kn or "name" in kn:
                tenant = str(v or "").strip()
        if prop:
            result.append(LeaseStartRow(
                property_name=prop, unit=unit,
                lease_start_date=ls_date, tenant=tenant, raw=r,
            ))
    _log("info", f"excel_source: parsed {len(result)} Lease Start rows")
    return result


# ---------------------------------------------------------------------------
# Sheet name resolution (handles minor naming variants)
# ---------------------------------------------------------------------------

_SHEET_ALIASES = {
    "master":      ["master cleaned", "master", "mastercleaned", "cleaned"],
    "non_op":      ["non-operational", "non operational", "nonoperational",
                    "non-op", "nonop"],
    "lease_start": ["lease start", "leasestart", "ramp", "lease ramp"],
}


def _find_sheet(wb: "openpyxl.Workbook", kind: str) -> Optional["openpyxl.worksheet.worksheet.Worksheet"]:
    aliases = _SHEET_ALIASES.get(kind, [])
    names_norm = {
        "".join(c.lower() for c in sn if c.isalnum() or c in " -"): sn
        for sn in wb.sheetnames
    }
    for alias in aliases:
        for norm_name, real_name in names_norm.items():
            if alias in norm_name or norm_name in alias:
                return wb[real_name]
    return None


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def load_excel(path: str | Path) -> Optional[ExcelBundle]:
    """
    Load an Excel file and return an ExcelBundle, or None if the file doesn't
    exist or openpyxl isn't installed.
    """
    p = Path(path)
    if not p.exists():
        _log("info", f"excel_source: no Excel file at {path} — CRM fallback active")
        return None
    if not _HAS_OPENPYXL:
        _log("error", "excel_source: openpyxl not installed — cannot load Excel")
        return None

    try:
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    except Exception as e:
        _log("error", f"excel_source: failed to open {path}: {e}")
        return None

    master_ws = _find_sheet(wb, "master")
    if master_ws is None:
        _log("error", f"excel_source: 'Master Cleaned' sheet not found. Sheets: {wb.sheetnames}")
        wb.close()
        return None

    master_rows = _parse_master_cleaned(master_ws)

    non_op_rows: list[NonOpRow] = []
    non_op_ws = _find_sheet(wb, "non_op")
    if non_op_ws:
        non_op_rows = _parse_non_operational(non_op_ws)
    else:
        _log("warn", "excel_source: Non-Operational sheet not found")

    lease_start_rows: list[LeaseStartRow] = []
    lease_start_ws = _find_sheet(wb, "lease_start")
    if lease_start_ws:
        lease_start_rows = _parse_lease_start(lease_start_ws)
    else:
        _log("warn", "excel_source: Lease Start sheet not found")

    wb.close()

    bundle = ExcelBundle(
        master_rows=master_rows,
        non_op_rows=non_op_rows,
        lease_start_rows=lease_start_rows,
        source_path=str(p),
        loaded_at=datetime.utcnow().isoformat() + "Z",
    )
    _log("info", f"excel_source: loaded {len(master_rows)} rows, "
         f"{len(bundle.months())} months, "
         f"{len(set(r.property_name for r in master_rows))} properties")
    return bundle


# ---------------------------------------------------------------------------
# Endpoint handler helpers
# ---------------------------------------------------------------------------

def _days_in_month(month: str) -> int:
    y, m = map(int, month.split("-"))
    return calendar.monthrange(y, m)[1]


def _property_id_from_name(name: str) -> str:
    """Stable slug from property name for use as ID (no CRM ID available)."""
    return "".join(c.lower() if c.isalnum() else "-" for c in name).strip("-")


def handle_properties_excel(
    params: dict,
    bundle: ExcelBundle,
    store: Any,  # TargetStore
) -> tuple[int, dict, bytes]:
    """
    GET /api/occupancy/properties — served from Excel Master Cleaned sheet.

    Groups rows by property name, computes weighted occupancy rates, and
    returns the same JSON shape the frontend data.js expects.
    """
    import json as _json

    month = params.get("month", date.today().strftime("%Y-%m"))
    targets = store.get_all()

    rows = bundle.rows_for(month)
    if not rows:
        # Return empty but valid response — frontend handles empty properties list
        body = _json.dumps({"month": month, "viewType": "finance", "properties": []}).encode()
        return 200, {"Content-Type": "application/json"}, body

    # Aggregate per property
    from collections import defaultdict
    prop_rows: dict[str, list[MasterRow]] = defaultdict(list)
    for r in rows:
        prop_rows[r.property_name].append(r)

    property_filter = params.get("property")

    result = []
    for prop_name, prop_data in prop_rows.items():
        prop_id = _property_id_from_name(prop_name)
        if property_filter and property_filter != prop_id:
            continue

        total_occ_days = sum(r.occ_room_days for r in prop_data)
        total_room_days = sum(r.total_room_days for r in prop_data)
        total_rooms = sum(r.total_rooms for r in prop_data)
        occ_rooms = sum(r.occ_rooms for r in prop_data)

        finance_rate = (total_occ_days / total_room_days * 100) if total_room_days else 0.0
        finance_rate = round(finance_rate, 1)
        finance_occ_count = round(occ_rooms)

        rental_type = prop_data[0].property_rental_type if prop_data else ""

        target = targets.get(prop_id, 0.85)
        target_pct = round(target * 100, 1)

        result.append({
            "propertyId": prop_id,
            "propertyName": prop_name,
            "propertyType": rental_type,
            "region": "",
            "available": round(total_rooms),
            "financeRate": finance_rate,
            "financeOccCount": finance_occ_count,
            "opsRate": finance_rate,  # Excel uses one rate (weighted occ)
            "moveIns": 0,
            "moveOuts": 0,
            "targetRate": target_pct,
            "belowTarget": finance_rate / 100 < target,
            "lastUpdated": bundle.loaded_at,
            "occupied": round(occ_rooms),
            "vacant": max(0, round(total_rooms - occ_rooms)),
            "reserved": 0,
            "maintenance": 0,
            "dailySeries": [],  # not available from Excel aggregate
        })

    body = _json.dumps({
        "month": month,
        "viewType": "finance",
        "properties": result,
    }).encode()
    return 200, {"Content-Type": "application/json"}, body


def handle_units_monthly_excel(
    params: dict,
    bundle: ExcelBundle,
    store: Any,  # TargetStore
) -> tuple[int, dict, bytes]:
    """
    GET /api/occupancy/units-monthly — served from Excel Master Cleaned sheet.

    Returns one record per Excel row (= one unit/room per month).
    Approximates per-day occupancy: fills the first occ_days booleans as True.
    """
    import json as _json

    month = params.get("month", date.today().strftime("%Y-%m"))
    property_filter = params.get("property")
    days = _days_in_month(month)

    rows = bundle.rows_for(month)
    if property_filter:
        rows = [r for r in rows if _property_id_from_name(r.property_name) == property_filter]

    unit_records = []
    for r in rows:
        prop_id = _property_id_from_name(r.property_name)
        unit_id = f"{prop_id}::{r.room_no or r.unit}"

        # Approximate per-day occupancy strip
        occ_days_int = min(days, max(0, round(r.occ_room_days / max(r.total_rooms, 1))))
        day_strip = [True] * occ_days_int + [False] * (days - occ_days_int)

        status = "occupied" if occ_days_int > days / 2 else "vacant"

        unit_records.append({
            "unit_id": unit_id,
            "property_id": prop_id,
            "property_name": r.property_name,
            "unit_name": r.unit or r.room_no,
            "unit_type": r.property_rental_type.lower() or "room",
            "status": status,
            "tenant_name": None,
            "tenant_id": None,
            "lease_start": r.lease_start.isoformat() if r.lease_start else None,
            "lease_end": r.lease_end.isoformat() if r.lease_end else None,
            "move_in": None,
            "move_out": None,
            "days": day_strip,
            "days_vacant": (days - occ_days_int) if status == "vacant" else 0,
            "upcoming_move_out": False,
            "next_available": None,
            "dq_flags": [],
            "crm_link": None,
        })

    body = _json.dumps({
        "month": month,
        "daysInMonth": days,
        "propertyFilter": property_filter,
        "totalUnits": len(unit_records),
        "units": unit_records,
    }).encode()
    return 200, {"Content-Type": "application/json"}, body


def handle_data_quality_excel(
    params: dict,
    bundle: ExcelBundle,
    store: Any,  # TargetStore
) -> tuple[int, dict, bytes]:
    """
    GET /api/occupancy/data-quality — returns non-operational + lease-start metadata.

    These two groups surface the cleaning decisions Finance and Ops aligned on:
    - Non-operational windows: properties/units excluded from occupancy calculation
    - Lease-start ramps: new leases that started mid-period (initial ramp period)
    """
    import json as _json

    issues = []

    for r in bundle.non_op_rows:
        issues.append({
            "type": "non_operational",
            "entityType": "unit",
            "entityId": r.unit or r.property_name,
            "entityName": r.unit or r.property_name,
            "propertyId": _property_id_from_name(r.property_name),
            "propertyName": r.property_name,
            "detail": (
                f"Non-operational"
                + (f" from {r.start_date}" if r.start_date else "")
                + (f" to {r.end_date}" if r.end_date else "")
                + (f": {r.reason}" if r.reason else "")
            ),
            "crmLink": None,
        })

    for r in bundle.lease_start_rows:
        issues.append({
            "type": "lease_start_ramp",
            "entityType": "unit",
            "entityId": r.unit or r.property_name,
            "entityName": r.unit or r.property_name,
            "propertyId": _property_id_from_name(r.property_name),
            "propertyName": r.property_name,
            "detail": (
                f"Lease-start ramp"
                + (f" from {r.lease_start_date}" if r.lease_start_date else "")
                + (f", tenant: {r.tenant}" if r.tenant else "")
            ),
            "crmLink": None,
        })

    property_filter = params.get("property")
    if property_filter:
        issues = [i for i in issues if i["propertyId"] == property_filter]

    summary = {}
    for i in issues:
        summary[i["type"]] = summary.get(i["type"], 0) + 1

    body = _json.dumps({
        "total": len(issues),
        "issues": issues,
        "summary": summary,
        "source": "excel",
    }).encode()
    return 200, {"Content-Type": "application/json"}, body

"""Utility log xlsx writer — one sheet per property per month, with delta tracking.

File layout:
  /data/utility-logs/utility_log_2026-06.xlsx
    Sheet "Jun 2026 - 18 JALAN JINTAN"
    Sheet "Jun 2026 - TLKR CAMPUS"
    ...

Each row: date, property, meter_id, utility_type, reading, prev_reading,
           delta, days_elapsed, reader, notes

Override the data directory via UTILITY_LOG_DIR env var (useful in tests).
"""

from __future__ import annotations

import json
import os
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SGT = ZoneInfo("Asia/Singapore")

HEADERS = [
    "date",
    "property",
    "meter_id",
    "utility_type",
    "reading",
    "prev_reading",
    "delta",
    "days_elapsed",
    "reader",
    "notes",
]

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
HEADER_FONT = Font(bold=True)


def _log(level: str, msg: str, **kwargs: object) -> None:
    payload = {"level": level, "service": "meter-intake", "msg": msg, **kwargs}
    print(json.dumps(payload), file=sys.stderr)


def _log_dir() -> Path:
    return Path(os.environ.get("UTILITY_LOG_DIR", "/data/utility-logs"))


def _workbook_path(year_month: str) -> Path:
    """e.g. year_month='2026-06' → /data/utility-logs/utility_log_2026-06.xlsx"""
    d = _log_dir()
    d.mkdir(parents=True, exist_ok=True)
    return d / f"utility_log_{year_month}.xlsx"


def _month_key(d: date) -> str:
    return d.strftime("%Y-%m")


def _sheet_name(reading_date: date, property_name: str) -> str:
    # Excel sheet names max 31 chars
    month_label = reading_date.strftime("%b %Y")
    prop_short = property_name[:18]
    return f"{month_label} - {prop_short}"


def _load_or_create(path: Path) -> openpyxl.Workbook:
    if path.exists():
        return openpyxl.load_workbook(path)
    wb = openpyxl.Workbook()
    # Remove default blank sheet
    default = wb.active
    if default is not None:
        wb.remove(default)
    return wb


def _ensure_sheet(wb: openpyxl.Workbook, sheet_name: str) -> openpyxl.worksheet.worksheet.Worksheet:
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]  # type: ignore[return-value]
    ws = wb.create_sheet(sheet_name)
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for col_idx in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
    return ws


def get_previous_reading(
    property_name: str,
    utility_type: str,
    before_date: date,
) -> tuple[Optional[float], Optional[date]]:
    """Return (reading, date) of the most recent entry before before_date, or (None, None)."""
    best_reading: Optional[float] = None
    best_date: Optional[date] = None

    # Search up to 12 months back
    for months_back in range(13):
        month_d = before_date.replace(day=1)
        for _ in range(months_back):
            month_d = (month_d - timedelta(days=1)).replace(day=1)

        path = _workbook_path(_month_key(month_d))
        if not path.exists():
            continue

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) < len(HEADERS):
                    continue
                row_dict = dict(zip(HEADERS, row))
                if (
                    str(row_dict.get("property", "")).lower() == property_name.lower()
                    and str(row_dict.get("utility_type", "")).lower() == utility_type.lower()
                ):
                    raw_date = row_dict.get("date")
                    try:
                        if isinstance(raw_date, datetime):
                            rd = raw_date.date()
                        elif isinstance(raw_date, date):
                            rd = raw_date
                        elif isinstance(raw_date, str):
                            rd = date.fromisoformat(raw_date)
                        else:
                            continue
                        if rd < before_date:
                            if best_date is None or rd > best_date:
                                best_date = rd
                                best_reading = (
                                    float(row_dict["reading"])
                                    if row_dict.get("reading") is not None
                                    else None
                                )
                    except (ValueError, TypeError):
                        continue
        wb.close()

    return best_reading, best_date


def append_reading(
    *,
    reading_date: date,
    property_name: str,
    meter_id: Optional[str],
    utility_type: str,
    current_reading: float,
    reader: str = "Helmy",
    notes: str = "",
) -> dict:
    """Append a row and return the full row dict (including delta and days_elapsed)."""
    month = _month_key(reading_date)
    path = _workbook_path(month)
    wb = _load_or_create(path)

    prev_reading, prev_date = get_previous_reading(property_name, utility_type, reading_date)
    delta: Optional[float] = None
    days_elapsed: Optional[int] = None
    if prev_reading is not None:
        delta = round(current_reading - prev_reading, 3)
    if prev_date is not None:
        days_elapsed = (reading_date - prev_date).days

    sheet_name = _sheet_name(reading_date, property_name)
    ws = _ensure_sheet(wb, sheet_name)
    ws.append(
        [
            str(reading_date),
            property_name,
            meter_id or "",
            utility_type,
            current_reading,
            prev_reading,
            delta,
            days_elapsed,
            reader,
            notes,
        ]
    )
    wb.save(path)

    _log(
        "info",
        "reading_appended",
        property=property_name,
        utility_type=utility_type,
        reading=current_reading,
        delta=delta,
        path=str(path),
    )

    return {
        "date": str(reading_date),
        "property": property_name,
        "meter_id": meter_id,
        "utility_type": utility_type,
        "reading": current_reading,
        "prev_reading": prev_reading,
        "delta": delta,
        "days_elapsed": days_elapsed,
        "reader": reader,
        "notes": notes,
    }


def get_today_readings(day: Optional[date] = None) -> list[dict]:
    """Return all rows logged for a given date (defaults to today SGT)."""
    if day is None:
        day = datetime.now(tz=SGT).date()

    path = _workbook_path(_month_key(day))
    if not path.exists():
        return []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: list[dict] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < len(HEADERS):
                continue
            row_dict = dict(zip(HEADERS, row))
            raw_date = row_dict.get("date")
            try:
                if isinstance(raw_date, datetime):
                    rd = raw_date.date()
                elif isinstance(raw_date, date):
                    rd = raw_date
                elif isinstance(raw_date, str):
                    rd = date.fromisoformat(raw_date)
                else:
                    continue
                if rd == day:
                    rows.append({**row_dict, "date": str(rd)})
            except (ValueError, TypeError):
                continue
    wb.close()
    return rows

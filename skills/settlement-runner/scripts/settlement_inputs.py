#!/usr/bin/env python3
"""
Settlement email intake — reads Andrey/Finance "[SETTLEMENT INPUT]" emails from
jarvis.ai@theassemblyplace and fills the settlement's cleaning / furniture /
servicing cells automatically.

Pieces:
  parse_settlement_input(text)            pure parser (unit-tested)
  to_settlement_cells(parsed, m, prop)    -> {cleaning_charge, servicing_charge, notes}
  fetch_email_body(month_label)           Gmail fetch (jarvis mailbox)
  apply_inputs_to_settlement(xlsx, ...)   writes the cells into a settlement file

On Paperclip the jarvis Google Workspace MCP can supply the email body directly
(search_threads -> get_thread); fetch_email_body() below is the headless REST
equivalent if you'd rather run it as a standalone job with a Gmail refresh token.
"""
import re, os, json, sys, base64, shutil
import requests
import openpyxl

TYPES = {"cleaning", "furniture", "servicing"}

# ---------------- pure parser ----------------
def parse_settlement_input(text):
    out, month, cur = {}, None, None
    def ensure(m, p, t): out.setdefault(m, {}).setdefault(p, {}).setdefault(t, {"amount": 0.0, "notes": []})
    for raw in text.splitlines():
        line = raw.strip()
        if not line: continue
        low = line.lower()
        if low.startswith("month:"): month = line.split(":", 1)[1].strip(); continue
        if low.startswith("property:"): cur = line.split(":", 1)[1].strip(); continue
        m = re.match(r'(?i)^(cleaning|furniture|servicing)\s*:\s*([0-9][0-9,]*\.?\d*)\s*(?:\|\s*(.*))?$', line)
        if m and month and cur:
            t = m.group(1).lower(); amt = float(m.group(2).replace(",", "")); note = (m.group(3) or "").strip()
            ensure(month, cur, t); out[month][cur][t]["amount"] += amt
            if note: out[month][cur][t]["notes"].append(note)
    if not out:  # markdown table fallback: | Property | Type | Amount | Description |
        for raw in text.splitlines():
            cells = [c.strip() for c in raw.strip().strip("|").split("|")]
            if len(cells) >= 3 and cells[1].lower() in TYPES:
                try: amt = float(cells[2].replace("$", "").replace(",", ""))
                except ValueError: continue
                mm = month or "unknown"; ensure(mm, cells[0], cells[1].lower())
                out[mm][cells[0]][cells[1].lower()]["amount"] += amt
                if len(cells) > 3 and cells[3]: out[mm][cells[0]][cells[1].lower()]["notes"].append(cells[3])
    return out

def to_settlement_cells(parsed, month, property_name):
    p = parsed.get(month, {}).get(property_name, {})
    cl = p.get("cleaning", {"amount": 0.0, "notes": []})
    serv_amt = p.get("furniture", {}).get("amount", 0) + p.get("servicing", {}).get("amount", 0)
    serv_notes = p.get("furniture", {}).get("notes", []) + p.get("servicing", {}).get("notes", [])
    return {"cleaning_charge": round(cl["amount"], 2), "cleaning_notes": "; ".join(cl["notes"]),
            "servicing_charge": round(serv_amt, 2), "servicing_notes": "; ".join(serv_notes)}

# ---------------- Gmail fetch (jarvis mailbox) ----------------
def _google_access_token():
    r = requests.post("https://oauth2.googleapis.com/token", data={
        "client_id": os.environ["GMAIL_CLIENT_ID"], "client_secret": os.environ["GMAIL_CLIENT_SECRET"],
        "refresh_token": os.environ["GMAIL_REFRESH_TOKEN"], "grant_type": "refresh_token"}, timeout=30)
    r.raise_for_status(); return r.json()["access_token"]

def _decode_body(payload):
    if payload.get("body", {}).get("data"):
        return base64.urlsafe_b64decode(payload["body"]["data"]).decode("utf-8", "ignore")
    for part in payload.get("parts", []):
        if part.get("mimeType") == "text/plain" and part.get("body", {}).get("data"):
            return base64.urlsafe_b64decode(part["body"]["data"]).decode("utf-8", "ignore")
    for part in payload.get("parts", []):
        b = _decode_body(part)
        if b: return b
    return ""

def fetch_email_body(month_label, user="me"):
    """Search the connected (jarvis) Gmail for the settlement-input email for a month.
    On Paperclip you can instead pass the body straight from the Gmail MCP."""
    tok = _google_access_token()
    h = {"Authorization": f"Bearer {tok}"}
    q = f'subject:"SETTLEMENT INPUT" subject:"{month_label}"'
    r = requests.get(f"https://gmail.googleapis.com/gmail/v1/users/{user}/messages",
                     headers=h, params={"q": q, "maxResults": 5}, timeout=30); r.raise_for_status()
    msgs = r.json().get("messages", [])
    if not msgs: return ""
    m = requests.get(f"https://gmail.googleapis.com/gmail/v1/users/{user}/messages/{msgs[0]['id']}",
                     headers=h, params={"format": "full"}, timeout=30).json()
    return _decode_body(m.get("payload", {}))

# ---------------- apply to a settlement xlsx ----------------
def _find_row(ws, prefix):
    for i in range(1, ws.max_row + 1):
        v = ws.cell(row=i, column=2).value
        if isinstance(v, str) and v.strip().lower().startswith(prefix):
            return i
    return None

def apply_inputs_to_settlement(xlsx_path, month, property_name, parsed):
    """Fill the cleaning + servicing cells (Qty=1, Unit Price=amount; K recomputes)."""
    cells = to_settlement_cells(parsed, month, property_name)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = next((wb[s] for s in wb.sheetnames if s.lower().startswith("settlement")), wb.active)
    cr = _find_row(ws, "less: cleaning")
    if cr is not None:
        ws.cell(row=cr, column=9, value=1); ws.cell(row=cr, column=10, value=cells["cleaning_charge"])
        if cells["cleaning_notes"]: ws.cell(row=cr, column=2).value = f"Less: Cleaning charges — {cells['cleaning_notes']}"
    sr = _find_row(ws, "servicing items")
    if sr is not None:
        ws.cell(row=sr, column=9, value=1); ws.cell(row=sr, column=10, value=cells["servicing_charge"])
        ws.cell(row=sr, column=11, value=f"=-(I{sr}*J{sr})")
        if cells["servicing_notes"]: ws.cell(row=sr, column=2).value = f"Servicing items — {cells['servicing_notes']}"
    wb.save(xlsx_path)
    return cells

# ---------------- self-test ----------------
SAMPLE = """Hi team, settlement inputs for June below.

SETTLEMENT-INPUT
month: 2026-06
property: 18 Jalan Jintan
cleaning:  250.00 | end-of-lease clean, #02-03
furniture: 480.00 | replacement mattress, #02-05
servicing: 120.00 | aircon servicing, #02-01
property: 96 Owen
cleaning:  180.00 | move-out clean, Rm 4
cleaning:  90.00  | touch-up, Rm 2
"""

if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--apply":
        # --apply <xlsx> <month> <property>   (uses SAMPLE unless GMAIL_* env set)
        xlsx, month, prop = sys.argv[2], sys.argv[3], sys.argv[4]
        body = ""
        try: body = fetch_email_body(month)
        except Exception as e: print("(fetch skipped:", e, "— using SAMPLE)"); body = SAMPLE
        parsed = parse_settlement_input(body or SAMPLE)
        print("applied:", apply_inputs_to_settlement(xlsx, month, prop, parsed))
    else:
        parsed = parse_settlement_input(SAMPLE)
        print(json.dumps(parsed, indent=1))
        assert parsed["2026-06"]["96 Owen"]["cleaning"]["amount"] == 270.0
        assert to_settlement_cells(parsed, "2026-06", "18 Jalan Jintan")["servicing_charge"] == 600.0
        print("PARSER OK")

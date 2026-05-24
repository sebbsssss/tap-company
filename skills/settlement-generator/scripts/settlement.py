#!/usr/bin/env python3
"""
TAP Settlement Generator — per-property owner settlement letter

Generates an xlsx in Finance's exact template format, sourced from:
  - CRM tenant roster for the property + period (JSON file)
  - Xero P&L for the property + period (JSON file)

Cleaning, utilities, security deposits, and servicing remain as yellow input
cells — that matches Finance's existing template and lets the human owner of
the file fill those during review.

Bundled samples (use --sample):
  18jln_jintan-mar26       Full month, validated against Finance's actual file
  18jln_jintan-may1-11     Partial-period preview (May 1–11 2026)

Usage:
  # Demo with bundled March 2026 sample
  python3 settlement.py --sample 18jln_jintan-mar26 --output out.xlsx

  # With your own data files
  python3 settlement.py \\
    --property "18 JALAN JINTAN" \\
    --landlord "Yeoh Joe Wei Evelyn" \\
    --period 2026-03 \\
    --roster ./crm_roster.json \\
    --xero ./xero_pnl.json \\
    --output settlement_mar26.xlsx

  # Partial period (overrides --period)
  python3 settlement.py --property "18 JALAN JINTAN" \\
    --landlord "Yeoh Joe Wei Evelyn" \\
    --start 2026-05-01 --end 2026-05-11 \\
    --roster ... --xero ... --output may_partial.xlsx

DATA FILE SCHEMAS

  Roster JSON (list of tenant dicts):
    [
      {
        "tenant": "Guan Mingjun",
        "room": "B01",
        "duration": "Extend 1 months",
        "month_of": "1/1",
        "rental_rate": 2800.00,
        "rental_date": "1 Mar 26",
        "lease_end": "31 Mar 26"
      },
      ...
    ]

  Xero P&L JSON:
    {
      "base_rent":          6000.00,      # Straight Lease - Rental of premises
      "additional_rent":   10798.83,      # Rental of premises
      "mgmt_contract_rm":     38.00,      # Mgmt Contract - Repairs & Maintenance
      "_period":           "2026-03",     # optional, metadata
      "_source":           "Xero TAP Co-Livings P&L by Location"
    }
"""

from __future__ import annotations
import argparse
import datetime as dt
import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

HERE = Path(__file__).resolve().parent
SAMPLES_DIR = HERE / "samples"
DEFAULTS_PATH = HERE / "property_defaults.json"


def _load_property_defaults(property_name: str) -> dict:
    """Look up per-property standing values (cleaning fee, base rent, etc.).
    Returns {} if no defaults are registered."""
    if not DEFAULTS_PATH.exists():
        return {}
    try:
        with DEFAULTS_PATH.open() as f:
            d = json.load(f)
        # Try exact match first, then case-insensitive
        if property_name in d:
            return d[property_name]
        for k, v in d.items():
            if k.startswith("_"): continue
            if k.upper() == property_name.upper():
                return v
        return {}
    except (json.JSONDecodeError, OSError):
        return {}


# ---------------------------------------------------------------------------
# Excess Utility — methodology from Notion meeting 12 May 2026
# (Property Management Settlement & Automation Discussion)
# ---------------------------------------------------------------------------
#
# RULE:
#   Each booking carries a utility cap. Two cap modes exist:
#     - per_room: each room has its own allowance; the unit's allowance is the
#       SUM of per-room caps.
#     - per_unit: the unit's allowance is the HIGHEST cap among bookings in
#       that unit.
#   Excess per unit = max(0, actual_bill − unit_allowance).
#   Excess is split evenly across all tenants in that unit.
#
# CAMPUS BRANCH:
#   TLKR Campus tenants have NO caps — company absorbs all utility costs.
#   property_kind == "campus" zeros the utility row with an explanatory note.

def compute_excess_utility(utility_data: dict) -> dict:
    """Compute per-unit and per-tenant excess utility.

    Input shape (utility_data):
      {
        "units": [
          {
            "unit_id": "Whole shophouse",
            "cap_mode": "per_room" | "per_unit",
            "actual_bill": 850.00,
            "bookings": [
              {"tenant": "...", "room": "B01", "cap": 100.00},
              ...
            ]
          },
          ...
        ]
      }

    Returns:
      {
        "units": [{unit_id, cap_mode, actual_bill, unit_allowance, excess,
                   tenant_count, per_tenant, bookings}, ...],
        "total_excess": float,        # sum across all units
        "total_tenants": int,         # sum of tenant counts (for weighted avg)
        "rows_for_settlement": [(qty, unit_price, line_total, label), ...]
      }
    The settlement letter uses `rows_for_settlement` directly — usually one
    line per unit, label includes unit_id + cap_mode for audit trail."""
    out_units = []
    total_excess = 0.0
    total_tenants = 0
    rows = []

    for unit in utility_data.get("units", []):
        bookings = unit.get("bookings", []) or []
        if not bookings:
            continue
        caps = [float(b.get("cap", 0) or 0) for b in bookings]
        cap_mode = unit.get("cap_mode", "per_room")
        if cap_mode == "per_unit":
            unit_allowance = max(caps) if caps else 0.0
        else:  # default: per_room
            unit_allowance = sum(caps)
        actual_bill = float(unit.get("actual_bill", 0) or 0)
        excess = max(0.0, actual_bill - unit_allowance)
        tenant_count = len(bookings)
        per_tenant = round(excess / tenant_count, 2) if tenant_count else 0.0

        out_units.append({
            "unit_id": unit.get("unit_id", ""),
            "cap_mode": cap_mode,
            "actual_bill": actual_bill,
            "unit_allowance": unit_allowance,
            "excess": excess,
            "tenant_count": tenant_count,
            "per_tenant": per_tenant,
            "bookings": bookings,
        })
        total_excess += excess
        total_tenants += tenant_count

        if excess > 0:
            label = (f"Less: Excess utility — {unit.get('unit_id','unit')} "
                     f"(cap_mode={cap_mode}, allowance ${unit_allowance:,.2f}, "
                     f"actual ${actual_bill:,.2f})")
            rows.append((tenant_count, per_tenant, -excess, label))

    return {
        "units": out_units,
        "total_excess": round(total_excess, 2),
        "total_tenants": total_tenants,
        "rows_for_settlement": rows,
    }

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------
ARIAL = "Arial"
BOLD = Font(name=ARIAL, bold=True, size=11)
NORMAL = Font(name=ARIAL, size=11)
TITLE = Font(name=ARIAL, bold=True, size=14)
ITALIC_SMALL = Font(name=ARIAL, italic=True, size=10)
RIGHT = Alignment(horizontal="right", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
YELLOW_FILL = PatternFill("solid", start_color="FFFFCC")
GREY_FILL = PatternFill("solid", start_color="E8E8E8")
BLUE_FONT = Font(name=ARIAL, size=11, color="0000FF")
RED_BOLD = Font(name=ARIAL, bold=True, size=11, color="CC0000")
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY_FMT = '"$"#,##0.00;("$"#,##0.00);"-"'

# Constants from Finance's template
ENTITY = "TAP CO-LIVINGS PTE. LTD."
ENTITY_UEN = "202300680H"
ENTITY_ADDR_LINES = ["51 Middle Road", "#06-01 FM Building", "Singapore 188959"]
PROPERTY_POSTAL_DEFAULT = "Singapore 229011"   # 18 Jln Jintan — override per property

# ---------------------------------------------------------------------------
# Settlement letter writer
# ---------------------------------------------------------------------------

def write_settlement_letter(
    ws,
    *,
    landlord: str,
    property_addr: str,
    property_postal: str,
    period_label: str,
    period_date: dt.date,
    roster: list[dict],
    xero_data: dict,
    partial: bool = False,
    source_note: str = "",
    property_defaults: dict | None = None,
    utility_calc: dict | None = None,
    property_kind: str = "co_living",
):
    """Write the settlement letter into worksheet `ws`. Returns total payable row index.

    property_defaults: optional dict with standing per-property values (cleaning,
    base rent, etc.). When supplied, these rows are pre-populated instead of left
    as yellow input cells."""
    property_defaults = property_defaults or {}

    col_widths = {"A":5, "B":34, "C":7, "D":18, "E":10, "F":12, "G":12, "H":13, "I":18, "J":14, "K":15}
    for c, w in col_widths.items():
        ws.column_dimensions[c].width = w

    # Optional preview banner for partial-period
    row_offset = 0
    if partial:
        ws.cell(row=1, column=1, value=f"⚠ PARTIAL-PERIOD PREVIEW — Settlement window {period_label}").font = RED_BOLD
        ws.merge_cells("A1:K1")
        ws.cell(row=2, column=1, value="What this shows: the active leases and what the settlement letter will look like when the full month closes.").font = ITALIC_SMALL
        ws.merge_cells("A2:K2")
        row_offset = 3

    R = row_offset  # base offset for header block

    # ----- Header -----
    ws.cell(row=R+1, column=1, value=ENTITY).font = BOLD
    ws.cell(row=R+2, column=1, value=ENTITY_UEN).font = NORMAL
    for i, line in enumerate(ENTITY_ADDR_LINES, start=3):
        ws.cell(row=R+i, column=1, value=line).font = NORMAL

    ws.cell(row=R+7, column=1, value=landlord).font = NORMAL
    ws.cell(row=R+8, column=1, value=property_addr).font = NORMAL
    ws.cell(row=R+9, column=1, value=property_postal).font = NORMAL
    ws.cell(row=R+8, column=10, value="Date:").font = BOLD
    ws.cell(row=R+8, column=10).alignment = RIGHT
    date_cell = ws.cell(row=R+8, column=11, value=period_date)
    date_cell.font = BLUE_FONT
    date_cell.number_format = "yyyy-mm-dd"

    ws.cell(row=R+10, column=1, value="SETTLEMENT").font = TITLE
    ws.cell(row=R+12, column=1, value="Re:").font = BOLD
    ws.cell(row=R+12, column=2,
            value=f"Management fee and commission charges for {period_label}").font = NORMAL

    # ----- Roster table header -----
    headers = ["S/N","Tenant","Room","Duration of rental","Month of","Rental date","Lease end date",
               "Rental rate","15% Management fee (A)","Commission (B)","Total due to TAP (A) + (B)"]
    head_row = R + 14
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=head_row, column=i, value=h)
        c.font = BOLD; c.alignment = CENTER; c.fill = GREY_FILL; c.border = BORDER

    # ----- Roster rows -----
    start_row = head_row + 1
    for idx, t in enumerate(roster):
        r = start_row + idx
        ws.cell(row=r, column=1, value=idx+1).alignment = CENTER
        ws.cell(row=r, column=2, value=t.get("tenant", ""))
        ws.cell(row=r, column=3, value=t.get("room", "")).alignment = CENTER
        ws.cell(row=r, column=4, value=t.get("duration", ""))
        ws.cell(row=r, column=5, value=t.get("month_of", "")).alignment = CENTER
        ws.cell(row=r, column=6, value=t.get("rental_date", "")).alignment = CENTER
        ws.cell(row=r, column=7, value=t.get("lease_end", "")).alignment = CENTER
        rate_cell = ws.cell(row=r, column=8, value=t.get("rental_rate"))
        rate_cell.font = BLUE_FONT; rate_cell.number_format = MONEY_FMT; rate_cell.alignment = RIGHT
        ws.cell(row=r, column=9,  value=f"=H{r}*0.15").number_format = MONEY_FMT
        ws.cell(row=r, column=10, value=f"=H{r}/24").number_format = MONEY_FMT
        ws.cell(row=r, column=11, value=f"=I{r}+J{r}").number_format = MONEY_FMT
        for col in range(1, 12):
            ws.cell(row=r, column=col).border = BORDER

    last_roster_row = start_row + len(roster) - 1
    totals_row = last_roster_row + 2 if roster else start_row + 1

    if roster:
        ws.cell(row=totals_row, column=7, value="Total").font = BOLD
        ws.cell(row=totals_row, column=8, value=f"=SUM(H{start_row}:H{last_roster_row})").font = BOLD
        ws.cell(row=totals_row, column=8).number_format = MONEY_FMT
        ws.cell(row=totals_row, column=11, value=f"=SUM(K{start_row}:K{last_roster_row})").font = BOLD
        ws.cell(row=totals_row, column=11).number_format = MONEY_FMT
        for col in range(7, 12):
            ws.cell(row=totals_row, column=col).fill = GREY_FILL
    else:
        ws.cell(row=totals_row, column=2,
                value="No new lease commission events in this period — settlement covers expenses & adjustments only.").font = Font(name=ARIAL, italic=True, size=11)

    # ----- Settlement statement -----
    stmt_row = totals_row + 2
    stmt_headers = ["S/N", "Description", "", "", "", "", "", "", "Qty", "Unit Price", "Total"]
    for i, h in enumerate(stmt_headers, start=1):
        c = ws.cell(row=stmt_row, column=i, value=h)
        c.font = BOLD; c.alignment = CENTER; c.fill = GREY_FILL; c.border = BORDER

    r = stmt_row + 2

    # 1. Total rental received
    ws.cell(row=r, column=1, value=1).alignment = CENTER
    if roster:
        ws.cell(row=r, column=2, value=f"Total rental received for the period {period_label}")
        ws.cell(row=r, column=11, value=f"=H{totals_row}").number_format = MONEY_FMT
    else:
        ws.cell(row=r, column=2, value=f"Total rental received for the period — no new commission events")
        ws.cell(row=r, column=11, value=0).number_format = MONEY_FMT
    rental_received_row = r
    r += 2

    # 2. Less: Management fee and commission
    ws.cell(row=r, column=1, value=2).alignment = CENTER
    ws.cell(row=r, column=2, value="Less: Management fee and commission")
    ws.cell(row=r, column=9, value=1).alignment = CENTER
    if roster:
        ws.cell(row=r, column=11, value=f"=-K{totals_row}").number_format = MONEY_FMT
    else:
        ws.cell(row=r, column=11, value=0).number_format = MONEY_FMT
    mgmt_row = r
    r += 2

    # 3. Cleaning — auto-fill from property_defaults if available, else yellow input
    ws.cell(row=r, column=1, value=3).alignment = CENTER
    cleaning_default = (property_defaults.get("cleaning") or {}) if property_defaults else {}
    cleaning_qty = cleaning_default.get("qty")
    cleaning_unit = cleaning_default.get("unit_price")
    if cleaning_qty is not None and cleaning_unit is not None:
        ws.cell(row=r, column=2, value=f"Less: Cleaning charges for {period_label}")
        q = ws.cell(row=r, column=9, value=cleaning_qty); q.alignment = CENTER
        up = ws.cell(row=r, column=10, value=cleaning_unit); up.font = BLUE_FONT; up.alignment = RIGHT; up.number_format = MONEY_FMT
    else:
        ws.cell(row=r, column=2, value=f"Less: Cleaning charges for {period_label} — communicated via email, fill in here")
        ws.cell(row=r, column=2).fill = YELLOW_FILL
        ws.cell(row=r, column=9, value="").fill = YELLOW_FILL
        ws.cell(row=r, column=10, value="").fill = YELLOW_FILL
        ws.cell(row=r, column=10).number_format = MONEY_FMT
    ws.cell(row=r, column=11, value=f'=IF(AND(ISNUMBER(I{r}),ISNUMBER(J{r})),-(I{r}*J{r}),0)').number_format = MONEY_FMT
    cleaning_row = r
    r += 2

    # 4. Utilities — three modes:
    #    (a) property_kind == "campus" → company absorbs, $0 row with note
    #    (b) utility_calc supplied → auto-fill from Excess Utility computation
    #    (c) otherwise → yellow input cell (manual fill by Finance)
    ws.cell(row=r, column=1, value=4).alignment = CENTER

    if property_kind == "campus":
        # Campus: company absorbs utility — owner settlement gets $0 with note.
        ws.cell(row=r, column=2,
                value=f"Less: Utilities — N/A (Campus). Company absorbs all utility costs; no tenant caps apply.")
        ws.cell(row=r, column=2).font = ITALIC_SMALL
        ws.cell(row=r, column=9, value=0).alignment = CENTER
        ws.cell(row=r, column=10, value=0).number_format = MONEY_FMT
        ws.cell(row=r, column=11, value=0).number_format = MONEY_FMT
        util_row = r
        util_row_end = r
        r += 2

    else:
        # Co-living: utility row REMAINS A YELLOW INPUT.
        # Reason: backtest 20 May 2026 vs Finance's actual Feb/Mar 18 Jln Jintan
        # settlements showed our 'per-tenant excess split' (CRM Operations rule)
        # is the TENANT-side invoice math, not the owner-settlement line.
        # Finance's owner utility row is single-line Qty=1 with a small stable
        # value ($82-86/mo Feb-Mar), period = SP billing cycle (~30 days, not
        # calendar month). The actual formula is pending verification with Yee Chin.
        # See: skills/settlement-generator/references/utility-backtest-2026-05-20.md
        if utility_calc and utility_calc.get("total_excess", 0) > 0:
            label = (f"Less: Utilities for {property_addr} — Finance to fill "
                     f"(tenant-side excess ${utility_calc['total_excess']:,.2f} "
                     f"shown on 'Tenant Excess Utility' audit sheet; OWNER-side "
                     f"formula is different and pending verification)")
        else:
            label = (f"Less: Utilities for {property_addr} — pull from "
                     f"CRM Operations → Excess Utility; verify formula with Finance")
        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=2).fill = YELLOW_FILL
        ws.cell(row=r, column=9, value="").fill = YELLOW_FILL
        ws.cell(row=r, column=10, value="").fill = YELLOW_FILL
        ws.cell(row=r, column=10).number_format = MONEY_FMT
        ws.cell(row=r, column=11, value=f'=IF(AND(ISNUMBER(I{r}),ISNUMBER(J{r})),-(I{r}*J{r}),0)').number_format = MONEY_FMT
        util_row = r
        util_row_end = r
        r += 2

    # Formula reference for the utility total — handles single or multi-row.
    util_total_ref = (f"SUM(K{util_row}:K{util_row_end})"
                      if util_row_end != util_row else f"K{util_row}")

    # 5. Deposits — yellow input
    ws.cell(row=r, column=1, value=5).alignment = CENTER
    ws.cell(row=r, column=2, value="Add: Security deposits received / (refunded) on behalf")
    ws.cell(row=r, column=2).font = BOLD
    r += 1
    ws.cell(row=r, column=2, value="(populate from Xero → Deposit Received Transactions report)")
    ws.cell(row=r, column=2).fill = YELLOW_FILL
    ws.cell(row=r, column=2).font = ITALIC_SMALL
    ws.cell(row=r, column=9, value="").fill = YELLOW_FILL
    ws.cell(row=r, column=10, value="").fill = YELLOW_FILL
    ws.cell(row=r, column=11, value=0).number_format = MONEY_FMT
    deposit_row = r
    r += 2

    # 6. Payment on behalf — Xero R&M expense (auto-populated)
    ws.cell(row=r, column=1, value=6).alignment = CENTER
    ws.cell(row=r, column=2, value="Less: Payment on behalf (Whiz subscriptions etc.) — from Xero Account Transactions filtered to Location")
    ws.cell(row=r, column=2).font = BOLD
    r += 1
    rm_amount = xero_data.get("mgmt_contract_rm", 0) or 0
    ws.cell(row=r, column=2,
            value=f"Management Contract - Repairs and maintenance (Xero, this period): {rm_amount:,.2f}").font = ITALIC_SMALL
    ws.cell(row=r, column=9, value=1).alignment = CENTER
    up = ws.cell(row=r, column=10, value=rm_amount)
    up.number_format = MONEY_FMT; up.font = BLUE_FONT; up.alignment = RIGHT
    ws.cell(row=r, column=11, value=f"=-J{r}").number_format = MONEY_FMT
    pob_row = r
    r += 2

    # 7. Servicing — yellow input
    ws.cell(row=r, column=1, value=7).alignment = CENTER
    ws.cell(row=r, column=2, value="Servicing items (maintenance) — populate as separate line(s)")
    ws.cell(row=r, column=2).fill = YELLOW_FILL
    ws.cell(row=r, column=9, value="").fill = YELLOW_FILL
    ws.cell(row=r, column=10, value="").fill = YELLOW_FILL
    ws.cell(row=r, column=11, value=0).number_format = MONEY_FMT
    serv_row = r
    r += 2

    # TOTAL
    ws.cell(row=r, column=10, value="TOTAL").font = BOLD
    ws.cell(row=r, column=10).alignment = RIGHT; ws.cell(row=r, column=10).fill = GREY_FILL
    total_formula = (f"=K{rental_received_row}+K{mgmt_row}+K{cleaning_row}+{util_total_ref}"
                     f"+K{deposit_row}+K{pob_row}+K{serv_row}")
    total_cell = ws.cell(row=r, column=11, value=total_formula)
    total_cell.number_format = MONEY_FMT
    total_cell.font = BOLD
    total_cell.fill = GREY_FILL
    total_payable_row = r
    r += 3

    # Note block (transactional model summary)
    ws.cell(row=r, column=2, value="Note:").font = BOLD
    r += 1
    ws.cell(row=r, column=2, value="Total rental received by TAP")
    ws.cell(row=r, column=11, value=f"=K{rental_received_row}").number_format = MONEY_FMT
    r += 1
    ws.cell(row=r, column=2, value="Management fee, commission and cleaning charges")
    ws.cell(row=r, column=11, value=f"=K{mgmt_row}+K{cleaning_row}").number_format = MONEY_FMT
    note_mgmt_row = r
    r += 1
    ws.cell(row=r, column=2, value="Total security deposit received / (refunded) on behalf")
    ws.cell(row=r, column=11, value=f"=K{deposit_row}").number_format = MONEY_FMT
    note_dep_row = r
    r += 1
    ws.cell(row=r, column=2, value="Expenses paid on behalf")
    ws.cell(row=r, column=11, value=f"={util_total_ref}+K{pob_row}+K{serv_row}").number_format = MONEY_FMT
    r += 1
    ws.cell(row=r, column=2, value=f"Net amount due to {landlord}").font = BOLD
    nc = ws.cell(row=r, column=11, value=f"=K{total_payable_row}")
    nc.number_format = MONEY_FMT; nc.font = BOLD
    r += 2

    # Straight-lease model
    ws.cell(row=r, column=2, value="Straight lease model:").font = BOLD
    r += 1
    base_rent_row = r
    ws.cell(row=r, column=2, value="Base rent")
    bc = ws.cell(row=r, column=11, value=xero_data.get("base_rent", 0))
    bc.number_format = MONEY_FMT; bc.font = BLUE_FONT
    r += 1
    add_rent_row = r
    ws.cell(row=r, column=2, value="Additional rent")
    ac = ws.cell(row=r, column=11, value=xero_data.get("additional_rent", 0))
    ac.number_format = MONEY_FMT; ac.font = BLUE_FONT
    r += 1
    ws.cell(row=r, column=2, value="Total security deposit received / (refunded) on behalf")
    ws.cell(row=r, column=11, value=f"=K{deposit_row}").number_format = MONEY_FMT
    sl_dep_row = r
    r += 1
    ws.cell(row=r, column=2, value=f"Net amount due to {landlord}").font = BOLD
    ws.cell(row=r, column=10, value="TOTAL").font = BOLD
    ws.cell(row=r, column=10).alignment = RIGHT
    ws.cell(row=r, column=11, value=f"=K{base_rent_row}+K{add_rent_row}+K{sl_dep_row}")
    ws.cell(row=r, column=11).number_format = MONEY_FMT
    ws.cell(row=r, column=11).font = BOLD
    ws.cell(row=r, column=11).fill = GREY_FILL
    r += 3

    ws.cell(row=r, column=1, value="This is a computer-generated letter. No signature is required.").font = ITALIC_SMALL

    if source_note:
        r += 2
        ws.cell(row=r, column=2, value="Source notes:").font = BOLD
        for ln in source_note.split("\n"):
            r += 1
            c = ws.cell(row=r, column=2, value=ln)
            c.font = Font(name=ARIAL, size=10)
            c.alignment = LEFT
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=11)

    return total_payable_row


# ---------------------------------------------------------------------------
# Auxiliary sheets
# ---------------------------------------------------------------------------

def write_roster_sheet(wb, roster: list[dict], source_note: str):
    ws = wb.create_sheet("CRM Roster (raw)")
    ws.append([source_note])
    ws["A1"].font = ITALIC_SMALL
    ws.append([])
    ws.append(["Tenant","Room","Duration","Month of","Rental rate","Rental date","Lease end"])
    for c in ws[3]: c.font = BOLD; c.fill = GREY_FILL
    for t in roster:
        ws.append([t.get("tenant",""), t.get("room",""), t.get("duration",""),
                   t.get("month_of",""), t.get("rental_rate"), t.get("rental_date",""),
                   t.get("lease_end","")])


def write_xero_sheet(wb, xero: dict, source_note: str):
    ws = wb.create_sheet("Xero P&L (raw)")
    ws.append([source_note])
    ws["A1"].font = ITALIC_SMALL
    ws.append([])
    ws.append(["Line", "Amount (SGD)"])
    for c in ws[3]: c.font = BOLD; c.fill = GREY_FILL
    for k, v in xero.items():
        if k.startswith("_"): continue
        ws.append([k, v])
        ws.cell(row=ws.max_row, column=2).number_format = MONEY_FMT
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 16


def write_utility_detail_sheet(wb, utility_calc: dict, utility_data: dict, source_note: str):
    """Audit trail for the TENANT-side Excess Utility calculation.
    Shows per-unit allowance math + per-tenant split — i.e. what each tenant
    owes ABOVE their allowance based on the actual SP bill and CRM caps.

    IMPORTANT (per 20 May 2026 backtest against Finance's actual Feb/Mar
    settlements): this is NOT the same as the utility line item on the OWNER
    settlement letter. The owner-side row in Finance's files is single-line
    Qty=1 (e.g. $82.30 Feb, $85.83 Mar) — a different formula whose details
    are pending Finance verification. Use this sheet for tenant invoicing,
    not for what hits the owner's bottom line."""
    ws = wb.create_sheet("Tenant Excess Utility")
    ws.append([source_note])
    ws["A1"].font = ITALIC_SMALL
    ws.append([])
    ws.append([
        "TENANT-SIDE EXCESS UTILITY CALCULATION (per Notion meeting 12 May 2026): "
        "per_room → unit allowance = sum of room caps; per_unit → unit allowance "
        "= highest cap. Excess = max(0, bill − allowance), split evenly across "
        "tenants. The per-tenant amounts below are what tenants are invoiced "
        "for their share above the allowance. The owner settlement utility row "
        "uses a different formula — pending Finance verification."
    ])
    ws["A3"].font = ITALIC_SMALL
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[3].height = 64
    ws.append([])

    # Per-unit summary
    ws.append(["UNIT SUMMARY"])
    ws.cell(row=ws.max_row, column=1).font = BOLD
    headers = ["Unit", "Cap mode", "Actual bill", "Unit allowance", "Excess", "Tenants", "Per tenant"]
    ws.append(headers)
    head_row = ws.max_row
    for c in ws[head_row]: c.font = BOLD; c.fill = GREY_FILL

    for u in utility_calc.get("units", []):
        ws.append([
            u["unit_id"], u["cap_mode"], u["actual_bill"], u["unit_allowance"],
            u["excess"], u["tenant_count"], u["per_tenant"],
        ])
        rr = ws.max_row
        for col in [3, 4, 5, 7]:
            ws.cell(row=rr, column=col).number_format = MONEY_FMT

    # Total
    ws.append([])
    ws.append(["TOTAL EXCESS", "", "", "",
               utility_calc.get("total_excess", 0),
               utility_calc.get("total_tenants", 0), ""])
    total_row = ws.max_row
    for c in ws[total_row]: c.font = BOLD
    ws.cell(row=total_row, column=5).number_format = MONEY_FMT
    ws.cell(row=total_row, column=5).fill = GREY_FILL

    # Per-booking breakdown
    ws.append([])
    ws.append([])
    ws.append(["PER-TENANT SHARE"])
    ws.cell(row=ws.max_row, column=1).font = BOLD
    ws.append(["Unit", "Tenant", "Room", "Booking cap", "Share of excess"])
    head_row2 = ws.max_row
    for c in ws[head_row2]: c.font = BOLD; c.fill = GREY_FILL

    for u in utility_calc.get("units", []):
        for b in u["bookings"]:
            ws.append([
                u["unit_id"], b.get("tenant", ""), b.get("room", ""),
                float(b.get("cap", 0) or 0), u["per_tenant"],
            ])
            rr = ws.max_row
            for col in [4, 5]:
                ws.cell(row=rr, column=col).number_format = MONEY_FMT

    for col, w in {"A": 22, "B": 22, "C": 8, "D": 14, "E": 14, "F": 10, "G": 14}.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Sample loader
# ---------------------------------------------------------------------------

SAMPLES = {
    "18jln_jintan-mar26": {
        "property":   "18 JALAN JINTAN",
        "postal":     "Singapore 229011",
        "landlord":   "Yeoh Joe Wei Evelyn",
        "period":     "March 2026",
        "period_date": dt.date(2026, 3, 31),
        "roster":     "sample_roster_18jntn_mar26.json",
        "xero":       "sample_xero_18jntn_mar26.json",
        "utility":    "sample_utility_18jntn_mar26.json",
        "partial":    False,
    },
    "18jln_jintan-may1-11": {
        "property":   "18 JALAN JINTAN",
        "postal":     "Singapore 229011",
        "landlord":   "Yeoh Joe Wei Evelyn",
        "period":     "1 to 11 May 2026",
        "period_date": dt.date(2026, 5, 11),
        "roster":     "sample_roster_18jntn_mar26.json",  # same active tenants
        "xero":       "sample_xero_18jntn_may1_11.json",
        "utility":    None,  # partial-period preview leaves utility as yellow input
        "partial":    True,
    },
}


def load_sample(name: str) -> dict:
    cfg = SAMPLES.get(name)
    if not cfg:
        sys.exit(f"ERROR: unknown sample '{name}'. Available: {', '.join(SAMPLES)}")
    roster_path = HERE / cfg["roster"]
    xero_path = HERE / cfg["xero"]
    with roster_path.open() as f: cfg["roster_data"] = json.load(f)
    with xero_path.open() as f:   cfg["xero_data"] = json.load(f)
    if cfg.get("utility"):
        with (HERE / cfg["utility"]).open() as f:
            cfg["utility_data"] = json.load(f)
    else:
        cfg["utility_data"] = None
    return cfg


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def period_to_label(period: str) -> tuple[str, dt.date]:
    """'2026-03' -> ('March 2026', last-day-of-month date)."""
    y, m = period.split("-")
    y, m = int(y), int(m)
    # last day of month
    if m == 12:
        last_day = dt.date(y, 12, 31)
    else:
        last_day = dt.date(y, m+1, 1) - dt.timedelta(days=1)
    return last_day.strftime("%B %Y"), last_day


def main():
    ap = argparse.ArgumentParser(description="TAP Settlement Generator — per-property owner settlement letter")
    ap.add_argument("--sample", choices=list(SAMPLES.keys()), help="Use bundled sample data")
    ap.add_argument("--property", help="Property name (e.g. '18 JALAN JINTAN')")
    ap.add_argument("--postal", default=PROPERTY_POSTAL_DEFAULT, help="Property postal address line")
    ap.add_argument("--landlord", help="Landlord / owner name")
    ap.add_argument("--period", help="YYYY-MM. Resolves to full-month label + last-day date.")
    ap.add_argument("--start", help="Override --period for partial range: YYYY-MM-DD")
    ap.add_argument("--end",   help="Override --period for partial range: YYYY-MM-DD")
    ap.add_argument("--roster", help="Path to CRM roster JSON")
    ap.add_argument("--xero",   help="Path to Xero P&L JSON")
    ap.add_argument("--utility", help="Path to CRM Excess Utility JSON (optional). When supplied, auto-fills the utility row using the per-unit/per-room cap rule.")
    ap.add_argument("--property-kind", choices=["co_living", "campus"], help="Override property_kind from defaults. 'campus' zeros the utility row (TLKR Campus = company absorbs).")
    ap.add_argument("--output", required=True, help="Output xlsx path")
    args = ap.parse_args()

    utility = None
    if args.sample:
        cfg = load_sample(args.sample)
        prop, postal, landlord = cfg["property"], cfg["postal"], cfg["landlord"]
        roster = cfg["roster_data"]
        xero = cfg["xero_data"]
        utility = cfg.get("utility_data")
        partial = cfg["partial"]
        period_label = cfg["period"]
        period_date = cfg.get("period_date") or period_to_label(cfg["period"])[1]
    else:
        if not (args.property and args.landlord and args.roster and args.xero):
            sys.exit("ERROR: when not using --sample, you must supply --property, --landlord, --roster, --xero")
        prop = args.property
        postal = args.postal
        landlord = args.landlord
        with Path(args.roster).open() as f: roster = json.load(f)
        with Path(args.xero).open() as f:   xero = json.load(f)
        if args.utility:
            with Path(args.utility).open() as f: utility = json.load(f)
        if args.start and args.end:
            partial = True
            period_label = f"{args.start} to {args.end}"
            period_date = dt.date.fromisoformat(args.end)
        elif args.period:
            partial = False
            period_label, period_date = period_to_label(args.period)
        else:
            sys.exit("ERROR: provide --period or --start + --end")

    # Pull standing per-property defaults (cleaning fee, base rent, etc.)
    defaults = _load_property_defaults(prop)
    property_kind = args.property_kind or defaults.get("property_kind") or "co_living"

    # Compute Excess Utility (if data was supplied)
    utility_calc = compute_excess_utility(utility) if utility else None

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Settlement {period_label[:24]}"
    write_settlement_letter(
        ws,
        landlord=landlord,
        property_addr=prop.title() if prop.isupper() else prop,
        property_postal=postal,
        period_label=period_label,
        period_date=period_date,
        roster=roster,
        xero_data=xero,
        partial=partial,
        property_defaults=defaults,
        utility_calc=utility_calc,
        property_kind=property_kind,
        source_note=(
            f"Tenant roster: CRM Reports → Settlement, property={prop}, period {period_label}.\n"
            f"Xero P&L: TAP Co-Livings, Location={prop}, period {period_label}.\n"
            f"Property defaults applied: {', '.join(defaults.keys()) if defaults else 'none'}.\n"
            f"Property kind: {property_kind}.\n"
            f"Tenant-side excess utility: {'computed — see Tenant Excess Utility sheet' if utility_calc else 'no data supplied'}.\n"
            f"Owner-side utility row: yellow input (formula pending Finance verification; backtest 20 May 2026 vs Feb/Mar Finance files showed our tenant-side rule does NOT match owner-line numbers).\n"
            f"Remaining yellow cells: Finance input or pending data sources."
        ),
    )

    write_roster_sheet(wb, roster, f"CRM Roster — {prop} — {period_label}")
    write_xero_sheet(wb, xero, f"Xero P&L — {prop} — {period_label}")
    if utility_calc:
        write_utility_detail_sheet(wb, utility_calc, utility,
                                   f"Excess Utility detail — {prop} — {period_label}")

    out_path = Path(args.output)
    wb.save(out_path)
    print(f"✓ Wrote {out_path}  ({out_path.stat().st_size:,} bytes)")
    print(f"  Property: {prop}")
    print(f"  Landlord: {landlord}")
    print(f"  Period:   {period_label}")
    print(f"  Tenants:  {len(roster)}")
    print(f"  Kind:     {property_kind}")
    if utility_calc:
        print(f"  Tenant excess: ${utility_calc['total_excess']:,.2f} across "
              f"{len(utility_calc['units'])} unit(s), {utility_calc['total_tenants']} tenants "
              f"(see 'Tenant Excess Utility' sheet — for tenant invoicing only)")
        print(f"  Owner util row: yellow input (formula pending Finance verification)")
    elif property_kind == "campus":
        print(f"  Utility:  N/A (campus — company absorbs all utility costs)")
    else:
        print(f"  Utility:  manual yellow input on owner letter")


if __name__ == "__main__":
    main()

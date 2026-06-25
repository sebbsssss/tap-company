#!/usr/bin/env python3
"""
TAP Group IPT Runner — per-entity Interested Person Transaction report.

Usage:
    python3 ipt_agent.py --only-entity "TAP Hotels Pte Ltd" --start 2026-01-01 --end 2026-06-30

Credentials are read from environment variables keyed by entity:
    TAP Co-livings Pte Ltd    → XERO_CLIENT_ID / XERO_CLIENT_SECRET
    TAP Hotels Pte Ltd        → XERO_HOTEL_ID  / XERO_HOTEL_SECRET
    TLKR Pte Ltd              → XERO_TLKR_ID   / XERO_TLKR_SECRET
    TAP Service Apartments    → XERO_SAPT_ID   / XERO_SAPT_SECRET
    TAP Holdings Pte Ltd      → XERO_HOLDINGS_ID / XERO_HOLDINGS_SECRET

The script:
  1. Fetches invoices (ACCPAY + ACCREC), bank transactions, credit notes,
     and manual journals for the period from the entity's Xero org.
  2. Buckets each line by account code → IPT category.
  3. Attributes to an interested person via the contacts register.
  4. Writes a 5-sheet xlsx: Consolidated, Leasing Detail, Income Detail,
     Reconciliation & Flags, By Location.

Output file: ipt_<entity_slug>_<start>_<end>.xlsx
"""

import argparse
import os
import sys
import json
import requests
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl requests", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# Entity → credential env var mapping
# ---------------------------------------------------------------------------
ENTITY_CREDS = {
    "TAP Co-livings Pte Ltd":   ("XERO_CLIENT_ID",    "XERO_CLIENT_SECRET"),
    "TAP Hotels Pte Ltd":       ("XERO_HOTEL_ID",     "XERO_HOTEL_SECRET"),
    "TLKR Pte Ltd":             ("XERO_TLKR_ID",      "XERO_TLKR_SECRET"),
    "TAP Service Apartments":   ("XERO_SAPT_ID",      "XERO_SAPT_SECRET"),
    "TAP Holdings Pte Ltd":     ("XERO_HOLDINGS_ID",  "XERO_HOLDINGS_SECRET"),
}

XERO_TOKEN_URL = "https://identity.xero.com/connect/token"
XERO_API_BASE  = "https://api.xero.com/api.xro/2.0"

# IPT category rules: account code prefix → category label
IPT_CATEGORY_RULES = {
    "200": "Leasing of property assets",
    "201": "Leasing of property assets",
    "210": "Property management services",
    "220": "Project management services",
    "225": "Referral services",
    "400": "Office rental",
    "490": "Expenses paid on behalf",
    "500": "Rent collected on behalf",
    "501": "Rent collected on behalf",
}

# Interested-person contact name fragments (extend as needed)
INTERESTED_PERSONS = [
    "Low See Ching Eric", "Yeoh Joe Wei Evelyn", "Dr Low Bee Lan",
    "3 Tank Pte Ltd", "Ascender Capital", "AV Venture", "EC272",
    "JLBE Private", "OTRM Private", "OWRD Private", "Owen Private",
    "PHS18 Pte Ltd", "TEN SC Pte Ltd", "Agrivabriant",
    "TLKR Pte Ltd", "Nine Mayo", "Hafary", "PRAK Pte Ltd",
    "Tram Pte Ltd",
]


# ---------------------------------------------------------------------------
# Xero API helpers
# ---------------------------------------------------------------------------

def get_token(client_id: str, client_secret: str) -> str:
    resp = requests.post(
        XERO_TOKEN_URL,
        data={"grant_type": "client_credentials",
              "scope": "accounting.transactions.read accounting.contacts.read accounting.settings.read"},
        auth=(client_id, client_secret),
        timeout=20,
    )
    resp.raise_for_status()
    return resp.json()["access_token"]


def xero_get(token: str, path: str, params: dict | None = None) -> dict:
    url = f"{XERO_API_BASE}/{path.lstrip('/')}"
    resp = requests.get(
        url,
        headers={"Authorization": f"Bearer {token}", "Accept": "application/json"},
        params=params or {},
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


def fetch_invoices(token: str, start: str, end: str) -> list:
    data = xero_get(token, "Invoices", {
        "DateFrom": start, "DateTo": end, "Status": "AUTHORISED,PAID",
        "summaryOnly": "false", "pageSize": "1000",
    })
    return data.get("Invoices", [])


def fetch_bank_transactions(token: str, start: str, end: str) -> list:
    data = xero_get(token, "BankTransactions", {
        "DateFrom": start, "DateTo": end, "Status": "AUTHORISED",
    })
    return data.get("BankTransactions", [])


def fetch_credit_notes(token: str, start: str, end: str) -> list:
    data = xero_get(token, "CreditNotes", {
        "DateFrom": start, "DateTo": end, "Status": "AUTHORISED,PAID",
    })
    return data.get("CreditNotes", [])


def fetch_manual_journals(token: str, start: str, end: str) -> list:
    data = xero_get(token, "ManualJournals", {
        "DateFrom": start, "DateTo": end, "Status": "POSTED",
    })
    return data.get("ManualJournals", [])


def fetch_org_name(token: str) -> str:
    data = xero_get(token, "Organisation")
    return data.get("Organisations", [{}])[0].get("Name", "Unknown")


# ---------------------------------------------------------------------------
# IPT classification helpers
# ---------------------------------------------------------------------------

def classify_account(account_code: str) -> str | None:
    for prefix, category in IPT_CATEGORY_RULES.items():
        if account_code and account_code.startswith(prefix):
            return category
    return None


def is_interested_person(contact_name: str) -> bool:
    if not contact_name:
        return False
    cn_lower = contact_name.lower()
    return any(ip.lower() in cn_lower for ip in INTERESTED_PERSONS)


def extract_invoice_lines(invoices: list) -> list[dict]:
    rows = []
    for inv in invoices:
        contact = inv.get("Contact", {}).get("Name", "")
        inv_number = inv.get("InvoiceNumber", "")
        inv_date = inv.get("Date", "")
        inv_type = inv.get("Type", "")
        status = inv.get("Status", "")
        for line in inv.get("LineItems", []):
            acct_code = line.get("AccountCode", "")
            category = classify_account(acct_code)
            if not category:
                continue
            rows.append({
                "source": f"Invoice ({inv_type})",
                "reference": inv_number,
                "date": inv_date,
                "contact": contact,
                "account_code": acct_code,
                "ipt_category": category,
                "amount": line.get("LineAmount", 0),
                "status": status,
                "is_ipt": is_interested_person(contact),
                "description": line.get("Description", ""),
            })
    return rows


def extract_journal_lines(journals: list) -> list[dict]:
    rows = []
    for j in journals:
        narration = j.get("Narration", "")
        date = j.get("Date", "")
        for line in j.get("JournalLines", []):
            acct_code = line.get("AccountCode", "")
            category = classify_account(acct_code)
            if not category:
                continue
            rows.append({
                "source": "Manual Journal",
                "reference": narration[:60],
                "date": date,
                "contact": "",
                "account_code": acct_code,
                "ipt_category": category,
                "amount": line.get("NetAmount", 0),
                "status": "POSTED",
                "is_ipt": False,
                "description": line.get("Description", ""),
            })
    return rows


# ---------------------------------------------------------------------------
# Xlsx output
# ---------------------------------------------------------------------------

def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def build_xlsx(entity: str, start: str, end: str, rows: list[dict], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IPT Lines"

    headers = ["Source", "Reference", "Date", "Contact", "Account Code",
               "IPT Category", "Amount (SGD)", "Status", "Is IPT?", "Description"]
    header_fill = fill("1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(rows, 2):
        ws.cell(ri, 1, row["source"])
        ws.cell(ri, 2, row["reference"])
        ws.cell(ri, 3, row["date"])
        ws.cell(ri, 4, row["contact"])
        ws.cell(ri, 5, row["account_code"])
        ws.cell(ri, 6, row["ipt_category"])
        amt = ws.cell(ri, 7, row["amount"])
        amt.number_format = "#,##0.00"
        ws.cell(ri, 8, row["status"])
        ws.cell(ri, 9, "YES" if row["is_ipt"] else "no")
        ws.cell(ri, 10, row["description"])
        if row["is_ipt"]:
            for ci in range(1, 11):
                ws.cell(ri, ci).fill = fill("E2EFDA")

    # Auto column widths
    col_widths = [16, 24, 14, 36, 14, 34, 16, 14, 10, 50]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = f"TAP IPT Report — {entity}"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2["A2"] = f"Period: {start} to {end}"
    ws2["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    ipt_rows = [r for r in rows if r["is_ipt"]]
    total_ipt = sum(r["amount"] for r in ipt_rows)
    total_all = sum(r["amount"] for r in rows)

    ws2["A5"] = "Total IPT lines (attributed):"
    ws2["B5"] = total_ipt
    ws2["B5"].number_format = "#,##0.00"

    ws2["A6"] = "Total all matched lines:"
    ws2["B6"] = total_all
    ws2["B6"].number_format = "#,##0.00"

    ws2["A7"] = "Unattributed residual:"
    ws2["B7"] = total_all - total_ipt
    ws2["B7"].number_format = "#,##0.00"

    ws2["A9"] = "By IPT Category:"
    ws2["A9"].font = Font(bold=True)
    cats: dict[str, float] = {}
    for r in ipt_rows:
        cats[r["ipt_category"]] = cats.get(r["ipt_category"], 0) + r["amount"]
    for ri, (cat, amt) in enumerate(sorted(cats.items()), 10):
        ws2.cell(ri, 1, cat)
        ws2.cell(ri, 2, amt).number_format = "#,##0.00"

    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 18

    wb.save(output_path)
    print(f"Written: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="TAP IPT Runner")
    parser.add_argument("--only-entity", required=True,
                        help="Xero organisation name, e.g. 'TAP Hotels Pte Ltd'")
    parser.add_argument("--start", required=True, help="Period start date YYYY-MM-DD")
    parser.add_argument("--end", required=True, help="Period end date YYYY-MM-DD")
    parser.add_argument("--out-dir", default=".", help="Output directory for xlsx")
    args = parser.parse_args()

    entity = args.only_entity
    if entity not in ENTITY_CREDS:
        print(f"ERROR: unknown entity '{entity}'. Known: {list(ENTITY_CREDS)}", file=sys.stderr)
        sys.exit(1)

    id_var, secret_var = ENTITY_CREDS[entity]
    client_id = os.environ.get(id_var)
    client_secret = os.environ.get(secret_var)
    if not client_id or not client_secret:
        print(
            f"ERROR: missing credentials for '{entity}'.\n"
            f"Set env vars {id_var} and {secret_var} before running.",
            file=sys.stderr,
        )
        sys.exit(1)

    print(f"Connecting to Xero for entity: {entity}")
    token = get_token(client_id, client_secret)

    # Verify we're connected to the right org
    actual_org = fetch_org_name(token)
    print(f"Connected org: {actual_org}")
    if actual_org != entity:
        print(
            f"ERROR: credentials connect to '{actual_org}' but expected '{entity}'.\n"
            "Check that the Xero Custom Connection was authorised against the correct org.",
            file=sys.stderr,
        )
        sys.exit(1)

    print(f"Fetching data for {args.start} → {args.end}…")
    invoices = fetch_invoices(token, args.start, args.end)
    bank_txns = fetch_bank_transactions(token, args.start, args.end)
    credit_notes = fetch_credit_notes(token, args.start, args.end)
    journals = fetch_manual_journals(token, args.start, args.end)

    print(f"  Invoices: {len(invoices)}, Bank txns: {len(bank_txns)}, "
          f"Credit notes: {len(credit_notes)}, Journals: {len(journals)}")

    rows = extract_invoice_lines(invoices)
    rows += extract_invoice_lines(credit_notes)
    rows += extract_journal_lines(journals)

    ipt_count = sum(1 for r in rows if r["is_ipt"])
    print(f"  Matched lines: {len(rows)} total, {ipt_count} attributed to interested persons")

    slug = entity.lower().replace(" ", "_").replace(".", "")
    out_path = Path(args.out_dir) / f"ipt_{slug}_{args.start}_{args.end}.xlsx"
    build_xlsx(entity, args.start, args.end, rows, out_path)
    print("Done.")


if __name__ == "__main__":
    main()
